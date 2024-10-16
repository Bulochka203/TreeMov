import datetime
import random
import calendar
from openpyxl import utils
from pprint import pprint

import openpyxl
from openpyxl.styles import Alignment

from .models import Attendance, StudentProfile

MONTHS = ['Январь', 'Февраль', 'Март', 'Апрель', 'Май', 'Июнь', 'Июль', 'Август', 'Сентярь', 'Октябрь', 'Ноябрь',
          'Декабрь']


def get_delay(debug=False):
    if debug:
        t = random.randint(30, 90)
        print(f'Следующий катаклизм наступит через {random.randint(30, 90)} секунд')
        return t * 10 ** 3

    negative_events_day_timer = random.randint(4, 7)  # через 4 - 7 дней
    negative_events_hour_timer = random.randint(8, 12)  # с 8 до 12 часов
    negative_events_minute_timer = random.randint(0, 59)  # с 0 до 59 минут

    now = datetime.date.today()
    datetime_now = datetime.datetime(now.year, now.month, now.day)
    event = datetime_now + datetime.timedelta(days=negative_events_day_timer,
                                              hours=negative_events_hour_timer,
                                              minutes=negative_events_minute_timer)

    delay = round((event - datetime.datetime.now()).total_seconds() * 1000)

    return delay


def creater_report(mentor_id, dates, groups):
    years_months_days = {}
    count_of_columns = {}

    for year in range(dates[0].year, dates[1].year + 1):
        count_of_columns[year] = 0
        years_months_days[year] = {}
        start_month, end_month = 1, 12

        if year == dates[0].year:
            start_month = dates[0].month - 1
            if dates[1].year == dates[0].year:
                end_month = dates[1].month
            else:
                end_month = 12
        elif year == dates[1].year:
            start_month = 0
            end_month = dates[1].month

        for month_num in range(start_month, end_month):
            first_day = 1
            _, last_day = calendar.monthrange(year, month_num + 1)

            if year == dates[0].year and month_num == dates[0].month - 1:
                first_day = dates[0].day
                if dates[1].year == dates[0].year and dates[1].month == dates[0].month:
                    last_day = dates[1].day
            elif year == dates[1].year and month_num == dates[1].month - 1:
                last_day = dates[1].day

            years_months_days[year][MONTHS[month_num]] = (first_day, last_day)
            count_of_columns[year] += last_day - first_day + 1

    groups_and_students = {group: [student for student in group.students.all()] for group in groups}
    pprint(groups_and_students, width=30, indent=4)
    print()
    pprint(years_months_days, width=30, indent=4)

    wb = openpyxl.Workbook()
    wb.remove(wb['Sheet'])
    for group, students in groups_and_students.items():
        wb.create_sheet(title=str(group))
        ws = wb[str(group)]

        year_cell_num = 2
        month_cell_num = 2
        days_cell_num = 2
        for year, month_dict in years_months_days.items():
            ws.merge_cells(start_column=year_cell_num, end_column=count_of_columns[year] + year_cell_num - 1,
                           start_row=1,
                           end_row=1)
            year_cell = ws.cell(row=1, column=year_cell_num, value=year)
            year_cell.alignment = Alignment(horizontal="center", vertical="center")
            year_cell_num += count_of_columns[year]
            for month, days in month_dict.items():
                ws.merge_cells(start_column=month_cell_num, end_column=month_cell_num + days[1] - days[0],
                               start_row=2,
                               end_row=2)
                month_cell = ws.cell(row=2, column=month_cell_num, value=month)
                month_cell.alignment = Alignment(horizontal="center", vertical="center")
                month_cell_num += days[1] - days[0] + 1
                for day in range(days[0], days[1] + 1):
                    day_cell = ws.cell(row=3, column=days_cell_num, value=day)
                    day_cell.alignment = Alignment(horizontal="center", vertical="center")
                    days_cell_num += 1
        for i in range(2, sum([count_of_columns[year] for year in count_of_columns.keys()]) + 2):
            idx_column = utils.cell.get_column_letter(i)
            ws.column_dimensions[idx_column].width = 3
            ws.column_dimensions[idx_column].height = 15

        ws.column_dimensions['A'].width = 30
        student_cell_num = 4
        for student in students:
            dates_of_visit = Attendance.objects.filter(student=student, group=group).values('dates')
            if dates_of_visit:
                dates_of_visit = list(dates_of_visit)[0]['dates']
                dates_of_visit.sort()
            else:
                dates_of_visit = []
            for d in dates_of_visit:
                if dates[0] <= d <= dates[1]:
                    plus_cell = ws.cell(row=student_cell_num, column=(d - dates[0]).days + 2, value='+')
                    plus_cell.alignment = Alignment(horizontal="center", vertical="center")

            student_cell = ws.cell(row=student_cell_num, column=1, value=str(student))
            student_cell.alignment = Alignment(horizontal="center", vertical="center")
            student_cell_num += 1

    path_to_file = f'reports/Отчёт{mentor_id}' \
                   f'{"".join([str(group) for group in groups_and_students.keys()])}' \
                   f'{str(datetime.datetime.now()).replace("-", "").replace(":", "").replace(" ", "")}.xlsx'
    wb.save(path_to_file)
    return path_to_file


def get_abilities(student_pk):
    student = StudentProfile.objects.get(pk=student_pk)
    tree = student.tree

